import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook, Workbook
from datetime import datetime
from tkinter import ttk

def save_information():
    program_name = entry_program_name.get()
    series_name = entry_series_name.get()
    program_number = entry_program_number.get()
    producer_name = producer_name_var.get()
    location = entry_location.get()
    address = entry_address.get()
    recording_type = recording_type_var.get()
    observations = text_observations.get("1.0", tk.END).strip()

    current_date = datetime.now().strftime("%Y-%m-%d")

    if program_name and location and address and producer_name and recording_type:
        try:
            try:
                workbook = load_workbook("producao.xlsx")
                sheet = workbook.active
            except:
                workbook = Workbook()
                sheet = workbook.active
                sheet.append(["Data", "Nome da Série", "Nome do Programa", "Número do Programa", "Nome do Produtor", "Local", "Endereço", "Tipo de Gravação", "Observações"])  # Headers

            sheet.append([current_date, series_name, program_name, program_number, producer_name, location, address, recording_type, observations])  # Data

            workbook.save("producao.xlsx")
            messagebox.showinfo("Sucesso", "Informações salvas com sucesso.")

            entry_program_name.delete(0, tk.END)
            entry_series_name.delete(0, tk.END)
            entry_program_number.delete(0, tk.END)
            producer_name_combobox.set("")
            entry_location.delete(0, tk.END)
            entry_address.delete(0, tk.END)
            recording_type_combobox.set("")
            text_observations.delete("1.0", tk.END)
        except Exception as e:
            messagebox.showerror("Erro", f"Ocorreu um erro: {e}")
    else:
        messagebox.showwarning("Aviso", "Preencha todos os campos obrigatórios.")

root = tk.Tk()
root.title("Inserir Informações de Produção")
root.geometry("800x600")  # Set the initial size of the window to half of the screen

label_series_name = tk.Label(root, text="Nome da Série:")
label_series_name.pack()

entry_series_name = tk.Entry(root, width=50)  # Increase the width of the entry
entry_series_name.pack()

label_program_name = tk.Label(root, text="Nome do Programa:")
label_program_name.pack()

entry_program_name = tk.Entry(root, width=50)  # Increase the width of the entry
entry_program_name.pack()

label_program_number = tk.Label(root, text="Número do Programa:")
label_program_number.pack()

entry_program_number = tk.Entry(root, width=50)  # Increase the width of the entry
entry_program_number.pack()

label_producer_name = tk.Label(root, text="Nome do Produtor:")
label_producer_name.pack()

producers = ["Bruno Barradas", "Monica Frazão", "Bia Correa"]
producer_name_var = tk.StringVar()
producer_name_combobox = ttk.Combobox(root, textvariable=producer_name_var, values=producers)
producer_name_combobox.pack()

label_location = tk.Label(root, text="Local:")
label_location.pack()

entry_location = tk.Entry(root, width=50)  # Increase the width of the entry
entry_location.pack()

label_address = tk.Label(root, text="Endereço:")
label_address.pack()

entry_address = tk.Entry(root, width=50)  # Increase the width of the entry
entry_address.pack()

label_recording_type = tk.Label(root, text="Tipo de Gravação:")
label_recording_type.pack()

recording_type_var = tk.StringVar()
recording_type_combobox = ttk.Combobox(root, textvariable=recording_type_var, values=["Externa", "Estúdio"])
recording_type_combobox.pack()

label_observations = tk.Label(root, text="Observações:")
label_observations.pack()

text_observations = tk.Text(root, height=5, width=50)  # Increase the width of the text field
text_observations.pack()

button_save = tk.Button(root, text="Salvar", command=save_information)
button_save.pack()

root.mainloop()